import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from config import REPORTS_DIR, create_application_directories
from database import fetch_all, fetch_one
from services.transaction_service import (
    format_currency,
    format_date_for_display,
)


REPORT_TYPES = (
    "Borçlu Müşteriler",
    "Borcu Kapanan Müşteriler",
    "Tüm Cari Hareketler",
    "Tahsilat Raporu",
    "Müşteri Ekstresi",
)


def parse_optional_date(
    date_text: str,
) -> tuple[bool, Optional[str], str]:
    """
    Boş bırakılabilen tarih alanını kontrol eder.

    Boşsa None döndürür.
    Doluysa GG.AA.YYYY biçimini YYYY-MM-DD biçimine çevirir.
    """

    date_text = date_text.strip()

    if not date_text:
        return True, None, ""

    try:
        parsed_date = datetime.strptime(
            date_text,
            "%d.%m.%Y",
        )

        return (
            True,
            parsed_date.strftime("%Y-%m-%d"),
            "",
        )

    except ValueError:
        return (
            False,
            None,
            "Tarihler GG.AA.YYYY biçiminde olmalıdır.",
        )


def validate_date_range(
    start_date_text: str,
    end_date_text: str,
) -> tuple[bool, Optional[str], Optional[str], str]:
    """
    Başlangıç ve bitiş tarihlerini kontrol eder.
    """

    start_success, start_date, start_message = (
        parse_optional_date(start_date_text)
    )

    if not start_success:
        return False, None, None, start_message

    end_success, end_date, end_message = (
        parse_optional_date(end_date_text)
    )

    if not end_success:
        return False, None, None, end_message

    if start_date and end_date and start_date > end_date:
        return (
            False,
            None,
            None,
            "Başlangıç tarihi bitiş tarihinden sonra olamaz.",
        )

    return True, start_date, end_date, ""


def build_date_filter(
    start_date: Optional[str],
    end_date: Optional[str],
    column_name: str = "transactions.transaction_date",
) -> tuple[str, list]:
    """
    SQL sorguları için tarih filtresi oluşturur.
    """

    conditions = []
    parameters = []

    if start_date:
        conditions.append(f"{column_name} >= ?")
        parameters.append(start_date)

    if end_date:
        conditions.append(f"{column_name} <= ?")
        parameters.append(end_date)

    if not conditions:
        return "", parameters

    return " AND " + " AND ".join(conditions), parameters


def get_customers_for_report() -> list[sqlite3.Row]:
    """
    Müşteri ekstresi seçimi için müşterileri getirir.
    """

    return fetch_all(
        """
        SELECT
            id,
            name,
            phone
        FROM customers
        ORDER BY name COLLATE NOCASE ASC
        """
    )


def get_customer_balance_report(
    only_debtors: bool,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> list[dict]:
    """
    Borçlu veya borcu kapanan müşterileri getirir.
    """

    date_filter, date_parameters = build_date_filter(
        start_date,
        end_date,
    )

    query = f"""
        SELECT
            customers.id,
            customers.name,
            customers.phone,

            COALESCE(
                SUM(
                    CASE
                        WHEN transactions.transaction_type = 'Borçlandırma'
                            THEN transactions.amount
                        ELSE 0
                    END
                ),
                0
            ) AS total_debt,

            COALESCE(
                SUM(
                    CASE
                        WHEN transactions.transaction_type = 'Ödeme'
                            THEN transactions.amount
                        ELSE 0
                    END
                ),
                0
            ) AS total_payment,

            COALESCE(
                SUM(
                    CASE
                        WHEN transactions.transaction_type = 'İndirim'
                            THEN transactions.amount
                        ELSE 0
                    END
                ),
                0
            ) AS total_discount,

            COALESCE(
                SUM(
                    CASE
                        WHEN transactions.transaction_type = 'İade'
                            THEN transactions.amount
                        ELSE 0
                    END
                ),
                0
            ) AS total_return

        FROM customers

        LEFT JOIN transactions
            ON transactions.customer_id = customers.id
            {date_filter}

        GROUP BY
            customers.id,
            customers.name,
            customers.phone

        ORDER BY customers.name COLLATE NOCASE ASC
    """

    rows = fetch_all(
        query,
        tuple(date_parameters),
    )

    results = []

    for row in rows:
        total_debt = float(row["total_debt"])
        total_payment = float(row["total_payment"])
        total_discount = float(row["total_discount"])
        total_return = float(row["total_return"])

        remaining = (
            total_debt
            - total_payment
            - total_discount
            - total_return
        )

        has_transaction = (
            total_debt
            + total_payment
            + total_discount
            + total_return
        ) > 0

        if only_debtors:
            should_include = remaining > 0.009
        else:
            should_include = (
                has_transaction
                and total_debt > 0
                and remaining <= 0.009
            )

        if should_include:
            results.append(
                {
                    "customer": row["name"],
                    "phone": row["phone"] or "-",
                    "total_debt": total_debt,
                    "total_payment": total_payment,
                    "total_discount": total_discount,
                    "total_return": total_return,
                    "remaining": max(remaining, 0.0),
                }
            )

    return results


def get_all_transactions_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> list[dict]:
    """
    Bütün cari hareketleri getirir.
    """

    date_filter, parameters = build_date_filter(
        start_date,
        end_date,
    )

    rows = fetch_all(
        f"""
        SELECT
            transactions.id,
            customers.name AS customer_name,
            transactions.transaction_type,
            transactions.amount,
            transactions.transaction_date,
            transactions.description
        FROM transactions
        INNER JOIN customers
            ON customers.id = transactions.customer_id
        WHERE 1 = 1
            {date_filter}
        ORDER BY
            transactions.transaction_date DESC,
            transactions.id DESC
        """,
        tuple(parameters),
    )

    return [
        {
            "id": row["id"],
            "customer": row["customer_name"],
            "transaction_type": row["transaction_type"],
            "amount": float(row["amount"]),
            "date": row["transaction_date"],
            "description": row["description"] or "-",
        }
        for row in rows
    ]


def get_payment_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> list[dict]:
    """
    Sadece ödeme işlemlerini getirir.
    """

    date_filter, parameters = build_date_filter(
        start_date,
        end_date,
    )

    rows = fetch_all(
        f"""
        SELECT
            transactions.id,
            customers.name AS customer_name,
            customers.phone,
            transactions.amount,
            transactions.transaction_date,
            transactions.description
        FROM transactions
        INNER JOIN customers
            ON customers.id = transactions.customer_id
        WHERE transactions.transaction_type = 'Ödeme'
            {date_filter}
        ORDER BY
            transactions.transaction_date DESC,
            transactions.id DESC
        """,
        tuple(parameters),
    )

    return [
        {
            "id": row["id"],
            "customer": row["customer_name"],
            "phone": row["phone"] or "-",
            "amount": float(row["amount"]),
            "date": row["transaction_date"],
            "description": row["description"] or "-",
        }
        for row in rows
    ]


def get_customer_statement(
    customer_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> tuple[bool, str, list[dict], dict]:
    """
    Seçilen müşterinin cari ekstresini getirir.
    """

    customer = fetch_one(
        """
        SELECT
            id,
            name,
            phone,
            address
        FROM customers
        WHERE id = ?
        """,
        (customer_id,),
    )

    if customer is None:
        return (
            False,
            "Müşteri bulunamadı.",
            [],
            {},
        )

    date_filter, date_parameters = build_date_filter(
        start_date,
        end_date,
    )

    parameters = [customer_id]
    parameters.extend(date_parameters)

    rows = fetch_all(
        f"""
        SELECT
            id,
            transaction_type,
            amount,
            transaction_date,
            description
        FROM transactions
        WHERE customer_id = ?
            {date_filter}
        ORDER BY
            transaction_date ASC,
            id ASC
        """,
        tuple(parameters),
    )

    statement_rows = []
    running_balance = 0.0

    total_debt = 0.0
    total_payment = 0.0
    total_discount = 0.0
    total_return = 0.0

    for row in rows:
        amount = float(row["amount"])
        transaction_type = row["transaction_type"]

        if transaction_type == "Borçlandırma":
            running_balance += amount
            total_debt += amount

        elif transaction_type == "Ödeme":
            running_balance -= amount
            total_payment += amount

        elif transaction_type == "İndirim":
            running_balance -= amount
            total_discount += amount

        elif transaction_type == "İade":
            running_balance -= amount
            total_return += amount

        statement_rows.append(
            {
                "id": row["id"],
                "date": row["transaction_date"],
                "transaction_type": transaction_type,
                "amount": amount,
                "description": row["description"] or "-",
                "balance": running_balance,
            }
        )

    summary = {
        "customer_name": customer["name"],
        "phone": customer["phone"] or "-",
        "address": customer["address"] or "-",
        "total_debt": total_debt,
        "total_payment": total_payment,
        "total_discount": total_discount,
        "total_return": total_return,
        "remaining": running_balance,
    }

    return True, "", statement_rows, summary


def generate_report(
    report_type: str,
    start_date_text: str = "",
    end_date_text: str = "",
    customer_id: Optional[int] = None,
) -> tuple[bool, str, list[str], list[list], dict]:
    """
    Seçilen rapor türüne göre tablo başlıklarını ve satırlarını oluşturur.
    """

    success, start_date, end_date, message = validate_date_range(
        start_date_text,
        end_date_text,
    )

    if not success:
        return False, message, [], [], {}

    if report_type == "Borçlu Müşteriler":
        records = get_customer_balance_report(
            only_debtors=True,
            start_date=start_date,
            end_date=end_date,
        )

        headers = [
            "Müşteri",
            "Telefon",
            "Toplam Borç",
            "Toplam Ödeme",
            "İndirim",
            "İade",
            "Kalan Borç",
        ]

        rows = [
            [
                record["customer"],
                record["phone"],
                format_currency(record["total_debt"]),
                format_currency(record["total_payment"]),
                format_currency(record["total_discount"]),
                format_currency(record["total_return"]),
                format_currency(record["remaining"]),
            ]
            for record in records
        ]

        return True, "", headers, rows, {}

    if report_type == "Borcu Kapanan Müşteriler":
        records = get_customer_balance_report(
            only_debtors=False,
            start_date=start_date,
            end_date=end_date,
        )

        headers = [
            "Müşteri",
            "Telefon",
            "Toplam Borç",
            "Toplam Ödeme",
            "İndirim",
            "İade",
            "Kalan Borç",
        ]

        rows = [
            [
                record["customer"],
                record["phone"],
                format_currency(record["total_debt"]),
                format_currency(record["total_payment"]),
                format_currency(record["total_discount"]),
                format_currency(record["total_return"]),
                format_currency(record["remaining"]),
            ]
            for record in records
        ]

        return True, "", headers, rows, {}

    if report_type == "Tüm Cari Hareketler":
        records = get_all_transactions_report(
            start_date=start_date,
            end_date=end_date,
        )

        headers = [
            "ID",
            "Müşteri",
            "İşlem Türü",
            "Tutar",
            "Tarih",
            "Açıklama",
        ]

        rows = [
            [
                record["id"],
                record["customer"],
                record["transaction_type"],
                format_currency(record["amount"]),
                format_date_for_display(record["date"]),
                record["description"],
            ]
            for record in records
        ]

        return True, "", headers, rows, {}

    if report_type == "Tahsilat Raporu":
        records = get_payment_report(
            start_date=start_date,
            end_date=end_date,
        )

        headers = [
            "ID",
            "Müşteri",
            "Telefon",
            "Ödeme Tutarı",
            "Tarih",
            "Açıklama",
        ]

        rows = [
            [
                record["id"],
                record["customer"],
                record["phone"],
                format_currency(record["amount"]),
                format_date_for_display(record["date"]),
                record["description"],
            ]
            for record in records
        ]

        total_payment = sum(
            record["amount"] for record in records
        )

        return (
            True,
            "",
            headers,
            rows,
            {
                "total_payment": total_payment,
            },
        )

    if report_type == "Müşteri Ekstresi":
        if customer_id is None:
            return (
                False,
                "Müşteri ekstresi için bir müşteri seçin.",
                [],
                [],
                {},
            )

        statement_success, statement_message, records, summary = (
            get_customer_statement(
                customer_id=customer_id,
                start_date=start_date,
                end_date=end_date,
            )
        )

        if not statement_success:
            return (
                False,
                statement_message,
                [],
                [],
                {},
            )

        headers = [
            "ID",
            "Tarih",
            "İşlem Türü",
            "Tutar",
            "Açıklama",
            "İşlem Sonrası Bakiye",
        ]

        rows = [
            [
                record["id"],
                format_date_for_display(record["date"]),
                record["transaction_type"],
                format_currency(record["amount"]),
                record["description"],
                format_currency(record["balance"]),
            ]
            for record in records
        ]

        return True, "", headers, rows, summary

    return False, "Geçersiz rapor türü.", [], [], {}


def create_safe_filename(text: str) -> str:
    """
    Dosya adına uygun güvenli metin oluşturur.
    """

    replacements = {
        "ç": "c",
        "Ç": "C",
        "ğ": "g",
        "Ğ": "G",
        "ı": "i",
        "İ": "I",
        "ö": "o",
        "Ö": "O",
        "ş": "s",
        "Ş": "S",
        "ü": "u",
        "Ü": "U",
        " ": "_",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    allowed_characters = set(
        "abcdefghijklmnopqrstuvwxyz"
        "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        "0123456789_-"
    )

    return "".join(
        character
        for character in text
        if character in allowed_characters
    )


def export_report_to_excel(
    report_title: str,
    headers: list[str],
    rows: list[list],
    summary: Optional[dict] = None,
) -> tuple[bool, str]:
    """
    Hazırlanan raporu biçimlendirilmiş Excel dosyası olarak kaydeder.
    """

    if not rows:
        return False, "Excel'e aktarılacak rapor verisi bulunmuyor."

    try:
        create_application_directories()

        filename = (
            f"{create_safe_filename(report_title)}_"
            f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        )

        file_path = REPORTS_DIR / filename

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Rapor"

        # Rapor başlığı
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(headers),
        )

        title_cell = worksheet.cell(
            row=1,
            column=1,
            value=report_title,
        )
        title_cell.font = Font(
            bold=True,
            size=16,
        )
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 25

        # Tablo başlıkları
        header_row = 3

        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )

            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Rapor satırları
        for row_index, row_data in enumerate(
            rows,
            start=header_row + 1,
        ):
            for column_index, value in enumerate(
                row_data,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        # Sütun genişliklerini güvenli şekilde ayarla.
        # Birleştirilmiş hücrelerden column_letter almıyoruz.
        for column_index in range(
            1,
            len(headers) + 1,
        ):
            column_letter = get_column_letter(
                column_index
            )
            max_length = 0

            for row_index in range(
                1,
                worksheet.max_row + 1,
            ):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                )

                cell_value = str(
                    cell.value or ""
                )

                max_length = max(
                    max_length,
                    len(cell_value),
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 3, 12),
                45,
            )

        # Özet bilgileri
        if summary:
            summary_start_row = (
                header_row
                + len(rows)
                + 3
            )

            customer_name = summary.get(
                "customer_name"
            )

            if customer_name:
                worksheet.cell(
                    row=summary_start_row,
                    column=1,
                    value="Müşteri",
                ).font = Font(bold=True)

                worksheet.cell(
                    row=summary_start_row,
                    column=2,
                    value=customer_name,
                )

                summary_start_row += 1

            summary_values = [
                (
                    "Toplam Borç",
                    summary.get("total_debt"),
                ),
                (
                    "Toplam Ödeme",
                    summary.get("total_payment"),
                ),
                (
                    "Toplam İndirim",
                    summary.get("total_discount"),
                ),
                (
                    "Toplam İade",
                    summary.get("total_return"),
                ),
                (
                    "Toplam Tahsilat",
                    summary.get("total_payment"),
                ),
                (
                    "Kalan Borç",
                    summary.get("remaining"),
                ),
            ]

            added_labels = set()

            for label, value in summary_values:
                if value is None:
                    continue

                if label in added_labels:
                    continue

                added_labels.add(label)

                worksheet.cell(
                    row=summary_start_row,
                    column=1,
                    value=label,
                ).font = Font(bold=True)

                worksheet.cell(
                    row=summary_start_row,
                    column=2,
                    value=format_currency(
                        float(value)
                    ),
                )

                summary_start_row += 1

        worksheet.freeze_panes = "A4"

        workbook.save(file_path)

        return (
            True,
            f"Excel raporu oluşturuldu:\n{file_path}",
        )

    except (
        OSError,
        ValueError,
        PermissionError,
    ) as error:
        return (
            False,
            (
                "Excel raporu oluşturulurken "
                f"hata oluştu: {error}"
            ),
        )

def register_pdf_font() -> str:
    """
    Türkçe karakter destekleyen Windows fontunu kaydeder.
    Font bulunamazsa Helvetica kullanır.
    """

    font_candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("C:/Windows/Fonts/segoeui.ttf"),
    ]

    for font_path in font_candidates:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(
                    TTFont(
                        "TurkishFont",
                        str(font_path),
                    )
                )

                return "TurkishFont"

            except Exception:
                continue

    return "Helvetica"


def export_report_to_pdf(
    report_title: str,
    headers: list[str],
    rows: list[list],
    summary: Optional[dict] = None,
) -> tuple[bool, str]:
    """
    Raporu PDF dosyası olarak kaydeder.
    """

    if not rows:
        return False, "PDF oluşturulacak rapor verisi bulunmuyor."

    try:
        create_application_directories()

        filename = (
            f"{create_safe_filename(report_title)}_"
            f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf"
        )

        file_path = REPORTS_DIR / filename
        font_name = register_pdf_font()

        document = SimpleDocTemplate(
            str(file_path),
            pagesize=landscape(A4),
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=17,
            alignment=TA_CENTER,
            spaceAfter=15,
        )

        normal_style = ParagraphStyle(
            name="TurkishNormal",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
        )

        content = [
            Paragraph(
                report_title,
                title_style,
            ),
            Spacer(1, 0.3 * cm),
        ]

        table_data = [
            [
                Paragraph(str(header), normal_style)
                for header in headers
            ]
        ]

        for row in rows:
            table_data.append(
                [
                    Paragraph(
                        str(value),
                        normal_style,
                    )
                    for value in row
                ]
            )

        available_width = landscape(A4)[0] - 2 * cm
        column_width = available_width / len(headers)

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[
                column_width
                for _ in headers
            ],
        )

        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#d9eaf7"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.black,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey,
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "ALIGN",
                        (0, 0),
                        (-1, 0),
                        "CENTER",
                    ),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [
                            colors.white,
                            colors.HexColor("#f4f4f4"),
                        ],
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                ]
            )
        )

        content.append(table)

        if summary:
            content.append(
                Spacer(
                    1,
                    0.5 * cm,
                )
            )

            summary_lines = []

            customer_name = summary.get("customer_name")

            if customer_name:
                summary_lines.append(
                    f"Müşteri: {customer_name}"
                )

            summary_values = [
                ("Toplam Borç", summary.get("total_debt")),
                ("Toplam Ödeme", summary.get("total_payment")),
                ("Toplam İndirim", summary.get("total_discount")),
                ("Toplam İade", summary.get("total_return")),
                ("Kalan Borç", summary.get("remaining")),
            ]

            for label, value in summary_values:
                if value is not None:
                    summary_lines.append(
                        f"{label}: {format_currency(float(value))}"
                    )

            for line in summary_lines:
                content.append(
                    Paragraph(
                        line,
                        normal_style,
                    )
                )

        document.build(content)

        return (
            True,
            f"PDF raporu oluşturuldu:\n{file_path}",
        )

    except Exception as error:
        return (
            False,
            f"PDF raporu oluşturulurken hata oluştu: {error}",
        )